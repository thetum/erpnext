import json
import subprocess
import urllib.request
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BASE_URL = "https://erp.chiangrai-nextstep.com"
API_KEY = "4739be7bcc70cb9"
API_SECRET = "5cee8032426f370"

def fetch_accounts():
    fields = '["name","account_name","parent_account","account_type","root_type","is_group","account_currency","company"]'
    from urllib.parse import quote
    url = f"{BASE_URL}/api/resource/Account?fields={quote(fields)}&limit=500&order_by=lft+asc"
    req = urllib.request.Request(url, headers={"Authorization": f"token {API_KEY}:{API_SECRET}"})
    with urllib.request.urlopen(req) as resp:
        data = json.loads(resp.read().decode())
    return data.get("data", [])

ROOT_TYPE_TH = {
    "Asset": "สินทรัพย์",
    "Liability": "หนี้สิน",
    "Equity": "ส่วนของผู้ถือหุ้น",
    "Income": "รายได้",
    "Expense": "ค่าใช้จ่าย",
}

ROOT_COLORS = {
    "Asset":    "DDEEFF",
    "Liability":"FFE5CC",
    "Equity":   "E2F0D9",
    "Income":   "E2EFDA",
    "Expense":  "FCE4D6",
}

GROUP_FILL = "F2F2F2"

HEADER_FILL  = "1F4E79"
HEADER_FONT  = Font(name="TH Sarabun New", bold=True, color="FFFFFF", size=13)
BODY_FONT    = Font(name="TH Sarabun New", size=12)
GROUP_FONT   = Font(name="TH Sarabun New", bold=True, size=12)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def build_tree(accounts):
    children = {}
    roots = []
    by_name = {a["name"]: a for a in accounts}
    for a in accounts:
        p = a.get("parent_account")
        if p:
            children.setdefault(p, []).append(a["name"])
        else:
            roots.append(a["name"])
    return roots, children, by_name

def flatten(roots, children, by_name, depth=0, result=None):
    if result is None:
        result = []
    for name in roots:
        a = by_name.get(name)
        if not a:
            continue
        result.append((a, depth))
        kids = children.get(name, [])
        if kids:
            flatten(kids, children, by_name, depth + 1, result)
    return result

def create_excel(accounts):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ผังบัญชี"
    ws.sheet_view.rightToLeft = False

    headers = ["รหัสบัญชี", "ชื่อบัญชี", "บัญชีแม่", "ประเภทราก", "ประเภทบัญชี", "สกุลเงิน", "กลุ่ม"]
    col_widths = [18, 45, 45, 18, 30, 10, 8]

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"ผังบัญชี (Chart of Accounts)  —  ส่งออก {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    title_cell.font = Font(name="TH Sarabun New", bold=True, size=16, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Header row
    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 22

    roots, children, by_name = build_tree(accounts)
    rows = flatten(roots, children, by_name)

    for row_idx, (acc, depth) in enumerate(rows, start=3):
        is_group = bool(acc.get("is_group"))
        root_type = acc.get("root_type", "")
        root_fill_color = ROOT_COLORS.get(root_type, "FFFFFF")

        indent = "  " * depth
        acct_name = indent + acc.get("account_name", "")
        # Extract account number prefix from name (e.g. "1110 - Cash - CN" -> "1110")
        raw_name = acc.get("name", "")
        parts = raw_name.split(" - ")
        acct_number = parts[0].strip() if parts else ""

        values = [
            acct_number,
            acct_name,
            acc.get("parent_account", "") or "",
            ROOT_TYPE_TH.get(root_type, root_type),
            acc.get("account_type", "") or "",
            acc.get("account_currency", "") or "",
            "ใช่" if is_group else "ไม่",
        ]

        fill_color = GROUP_FILL if is_group and depth > 0 else root_fill_color if depth == 0 else "FFFFFF"
        row_fill = PatternFill("solid", fgColor=fill_color)
        font = GROUP_FONT if is_group else BODY_FONT

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = font
            cell.fill = row_fill
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 2))

    # Freeze panes below header
    ws.freeze_panes = "A3"

    # Auto-filter on headers
    ws.auto_filter.ref = f"A2:G{2 + len(rows)}"

    ws.row_dimensions[1].height = 30
    for r in range(3, 3 + len(rows)):
        ws.row_dimensions[r].height = 18

    out_path = "/mnt/d/Job/erp-next/ผังบัญชี.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}  ({len(rows)} accounts)")

if __name__ == "__main__":
    accounts = fetch_accounts()
    print(f"Fetched {len(accounts)} accounts")
    create_excel(accounts)
